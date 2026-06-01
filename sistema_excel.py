import openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime
from pathlib import Path

class SistemaExcel:
    def __init__(self, archivo_excel="Libro1.xlsx"):
        self.archivo_excel = archivo_excel
        self.cargar_workbook()
    
    def cargar_workbook(self):
        """Carga el archivo Excel"""
        try:
            self.wb = openpyxl.load_workbook(self.archivo_excel)
        except FileNotFoundError:
            print(f"Error: No se encontró {self.archivo_excel}")
            self.wb = None
    
    def obtener_hoja(self, nombre_hoja):
        """Obtiene una hoja del workbook"""
        if self.wb and nombre_hoja in self.wb.sheetnames:
            return self.wb[nombre_hoja]
        return None
    
    def obtener_todas_hojas(self):
        """Retorna los nombres de todas las hojas"""
        if self.wb:
            return self.wb.sheetnames
        return []
    
    def agregar_ingreso(self, date, document_number, customer, category, payment_method, status, amount, notes=""):
        """Agrega una fila a la hoja Income"""
        ws = self.obtener_hoja("Income")
        if ws:
            nueva_fila = ws.max_row + 1
            ws[f'A{nueva_fila}'] = date
            ws[f'B{nueva_fila}'] = document_number
            ws[f'C{nueva_fila}'] = customer
            ws[f'D{nueva_fila}'] = category
            ws[f'E{nueva_fila}'] = payment_method
            ws[f'F{nueva_fila}'] = status
            ws[f'G{nueva_fila}'] = amount
            ws[f'H{nueva_fila}'] = notes
            self.guardar()
            return True
        return False
    
    def agregar_gasto(self, date, document_number, description, supplier, category, tipo, payment_method, status, amount, notes=""):
        """Agrega una fila a la hoja Expenses"""
        ws = self.obtener_hoja("Expenses")
        if ws:
            nueva_fila = ws.max_row + 1
            ws[f'A{nueva_fila}'] = date
            ws[f'B{nueva_fila}'] = document_number
            ws[f'C{nueva_fila}'] = description
            ws[f'D{nueva_fila}'] = supplier
            ws[f'E{nueva_fila}'] = category
            ws[f'F{nueva_fila}'] = tipo
            ws[f'G{nueva_fila}'] = payment_method
            ws[f'H{nueva_fila}'] = status
            ws[f'I{nueva_fila}'] = amount
            ws[f'J{nueva_fila}'] = notes
            self.guardar()
            return True
        return False
    
    def agregar_contrato_labor(self, date, document_number, contractor_name, description, payment_method, amount, notes=""):
        """Agrega una fila a la hoja Contract Labor"""
        ws = self.obtener_hoja("Contract Labor")
        if ws:
            nueva_fila = ws.max_row + 1
            ws[f'A{nueva_fila}'] = date
            ws[f'B{nueva_fila}'] = document_number
            ws[f'C{nueva_fila}'] = contractor_name
            ws[f'D{nueva_fila}'] = description
            ws[f'E{nueva_fila}'] = payment_method
            ws[f'F{nueva_fila}'] = amount
            ws[f'G{nueva_fila}'] = notes
            self.guardar()
            return True
        return False
    
    def obtener_ingresos(self):
        """Obtiene todos los ingresos"""
        ws = self.obtener_hoja("Income")
        datos = []
        if ws:
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0]:  # Si hay fecha
                    datos.append({
                        'Date': row[0],
                        'Document Number': row[1],
                        'Customer': row[2],
                        'Category': row[3],
                        'Payment Method': row[4],
                        'Status': row[5],
                        'Amount': row[6],
                        'Notes': row[7]
                    })
        return datos
    
    def obtener_gastos(self):
        """Obtiene todos los gastos"""
        ws = self.obtener_hoja("Expenses")
        datos = []
        if ws:
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0]:  # Si hay fecha
                    datos.append({
                        'Date': row[0],
                        'Document Number': row[1],
                        'Description': row[2],
                        'Supplier': row[3],
                        'Category': row[4],
                        'Type': row[5],
                        'Payment Method': row[6],
                        'Status': row[7],
                        'Amount': row[8],
                        'Notes': row[9]
                    })
        return datos
    
    def obtener_contratos_labor(self):
        """Obtiene todos los contratos de labor"""
        ws = self.obtener_hoja("Contract Labor")
        datos = []
        if ws:
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0]:  # Si hay fecha
                    datos.append({
                        'Date': row[0],
                        'Document Number': row[1],
                        'Contractor Name': row[2],
                        'Description': row[3],
                        'Payment Method': row[4],
                        'Amount': row[5],
                        'Notes': row[6]
                    })
        return datos
    
    def obtener_categorias_ingresos(self):
        """Obtiene todas las categorías de ingresos únicas"""
        ingresos = self.obtener_ingresos()
        categorias = set()
        for ingreso in ingresos:
            if ingreso['Category']:
                categorias.add(ingreso['Category'])
        return sorted(list(categorias))
    
    def obtener_categorias_gastos(self):
        """Obtiene todas las categorías de gastos únicas"""
        gastos = self.obtener_gastos()
        categorias = set()
        for gasto in gastos:
            if gasto['Category']:
                categorias.add(gasto['Category'])
        return sorted(list(categorias))
    
    def obtener_clientes(self):
        """Obtiene todos los clientes únicos"""
        ingresos = self.obtener_ingresos()
        clientes = set()
        for ingreso in ingresos:
            if ingreso['Customer']:
                clientes.add(ingreso['Customer'])
        return sorted(list(clientes))
    
    def obtener_proveedores(self):
        """Obtiene todos los proveedores únicos"""
        gastos = self.obtener_gastos()
        proveedores = set()
        for gasto in gastos:
            if gasto['Supplier']:
                proveedores.add(gasto['Supplier'])
        return sorted(list(proveedores))
    
    def obtener_metodos_pago(self):
        """Obtiene todos los métodos de pago únicos"""
        metodos = set()
        
        # De ingresos
        ingresos = self.obtener_ingresos()
        for ingreso in ingresos:
            if ingreso['Payment Method']:
                metodos.add(ingreso['Payment Method'])
        
        # De gastos
        gastos = self.obtener_gastos()
        for gasto in gastos:
            if gasto['Payment Method']:
                metodos.add(gasto['Payment Method'])
        
        return sorted(list(metodos))
    
    def calcular_totales_mes(self, año, mes):
        """Calcula totales de ingresos y gastos para un mes específico"""
        ingresos_total = 0
        gastos_total = 0
        
        # Ingresos
        ingresos = self.obtener_ingresos()
        for ingreso in ingresos:
            fecha = ingreso['Date']
            if fecha and isinstance(fecha, datetime):
                if fecha.year == año and fecha.month == mes:
                    if ingreso['Amount']:
                        ingresos_total += ingreso['Amount']
        
        # Gastos
        gastos = self.obtener_gastos()
        for gasto in gastos:
            fecha = gasto['Date']
            if fecha and isinstance(fecha, datetime):
                if fecha.year == año and fecha.month == mes:
                    if gasto['Amount']:
                        gastos_total += gasto['Amount']
        
        # Contratos de labor
        contratos = self.obtener_contratos_labor()
        for contrato in contratos:
            fecha = contrato['Date']
            if fecha and isinstance(fecha, datetime):
                if fecha.year == año and fecha.month == mes:
                    if contrato['Amount']:
                        gastos_total += contrato['Amount']
        
        return {
            'ingresos': ingresos_total,
            'gastos': gastos_total,
            'neto': ingresos_total - gastos_total
        }
    
    def guardar(self):
        """Guarda el workbook"""
        if self.wb:
            self.wb.save(self.archivo_excel)