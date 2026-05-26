import openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime
from pathlib import Path

class SistemaContabilidadExcel:
    def __init__(self, archivo_excel="Libro1.xlsx"):
        self.archivo_excel = archivo_excel
        self.wb = None
        self.cargar_libro()
    
    def cargar_libro(self):
        """Carga el libro Excel existente"""
        if Path(self.archivo_excel).exists():
            self.wb = openpyxl.load_workbook(self.archivo_excel)
        else:
            raise FileNotFoundError(f"El archivo {self.archivo_excel} no existe")
    
    def guardar_libro(self):
        """Guarda los cambios en el Excel"""
        self.wb.save(self.archivo_excel)
    
    def obtener_hoja(self, nombre_hoja):
        """Obtiene una hoja del libro"""
        if nombre_hoja in self.wb.sheetnames:
            return self.wb[nombre_hoja]
        return None
    
    def agregar_ingreso(self, fecha, documento, cliente, categoria, metodo_pago, estado, monto, notas):
        """Agrega un ingreso a la hoja Income"""
        ws = self.obtener_hoja("Income")
        if ws is None:
            return False
        
        # Encuentra la siguiente fila vacía
        row = 2
        while ws[f'A{row}'].value is not None:
            row += 1
        
        # Agrega los datos
        ws[f'A{row}'] = fecha  # Date
        ws[f'B{row}'] = documento  # Document Number
        ws[f'C{row}'] = cliente  # Customer
        ws[f'D{row}'] = categoria  # Category
        ws[f'E{row}'] = metodo_pago  # Payment Method
        ws[f'F{row}'] = estado  # Status
        ws[f'G{row}'] = monto  # Amount
        ws[f'H{row}'] = notas  # Notes
        
        self.guardar_libro()
        return True
    
    def agregar_gasto(self, fecha, documento, descripcion, proveedor, categoria, tipo, metodo_pago, estado, monto, notas):
        """Agrega un gasto a la hoja Expenses"""
        ws = self.obtener_hoja("Expenses")
        if ws is None:
            return False
        
        row = 2
        while ws[f'A{row}'].value is not None:
            row += 1
        
        ws[f'A{row}'] = fecha  # Date
        ws[f'B{row}'] = documento  # Document Number
        ws[f'C{row}'] = descripcion  # Description
        ws[f'D{row}'] = proveedor  # Supplier
        ws[f'E{row}'] = categoria  # Category
        ws[f'F{row}'] = tipo  # Type
        ws[f'G{row}'] = metodo_pago  # Payment Method
        ws[f'H{row}'] = estado  # Status
        ws[f'I{row}'] = monto  # Amount
        ws[f'J{row}'] = notas  # Notes
        
        self.guardar_libro()
        return True
    
    def agregar_contrato_labor(self, fecha, documento, contratista, descripcion, metodo_pago, monto, notas):
        """Agrega un pago a contratista en la hoja Contract Labor"""
        ws = self.obtener_hoja("Contract Labor")
        if ws is None:
            return False
        
        row = 5  # Comienza después de los encabezados
        while ws[f'A{row}'].value is not None:
            row += 1
        
        ws[f'A{row}'] = fecha  # Date
        ws[f'B{row}'] = documento  # Document #
        ws[f'C{row}'] = contratista  # Contractor Name
        ws[f'D{row}'] = descripcion  # Description
        ws[f'E{row}'] = metodo_pago  # Payment Method
        ws[f'F{row}'] = monto  # Amount
        ws[f'G{row}'] = notas  # Notes
        
        self.guardar_libro()
        return True
    
    def obtener_ingresos(self):
        """Obtiene todos los ingresos"""
        ws = self.obtener_hoja("Income")
        if ws is None:
            return []
        
        ingresos = []
        for row in range(2, ws.max_row + 1):
            if ws[f'A{row}'].value:
                ingresos.append({
                    'fecha': ws[f'A{row}'].value,
                    'documento': ws[f'B{row}'].value,
                    'cliente': ws[f'C{row}'].value,
                    'categoria': ws[f'D{row}'].value,
                    'metodo_pago': ws[f'E{row}'].value,
                    'estado': ws[f'F{row}'].value,
                    'monto': ws[f'G{row}'].value,
                    'notas': ws[f'H{row}'].value,
                })
        return ingresos
    
    def obtener_gastos(self):
        """Obtiene todos los gastos"""
        ws = self.obtener_hoja("Expenses")
        if ws is None:
            return []
        
        gastos = []
        for row in range(2, ws.max_row + 1):
            if ws[f'A{row}'].value:
                gastos.append({
                    'fecha': ws[f'A{row}'].value,
                    'documento': ws[f'B{row}'].value,
                    'descripcion': ws[f'C{row}'].value,
                    'proveedor': ws[f'D{row}'].value,
                    'categoria': ws[f'E{row}'].value,
                    'tipo': ws[f'F{row}'].value,
                    'metodo_pago': ws[f'G{row}'].value,
                    'estado': ws[f'H{row}'].value,
                    'monto': ws[f'I{row}'].value,
                    'notas': ws[f'J{row}'].value,
                })
        return gastos
    
    def obtener_contratos_labor(self):
        """Obtiene todos los pagos a contratistas"""
        ws = self.obtener_hoja("Contract Labor")
        if ws is None:
            return []
        
        contratos = []
        for row in range(5, ws.max_row + 1):
            if ws[f'A{row}'].value:
                contratos.append({
                    'fecha': ws[f'A{row}'].value,
                    'documento': ws[f'B{row}'].value,
                    'contratista': ws[f'C{row}'].value,
                    'descripcion': ws[f'D{row}'].value,
                    'metodo_pago': ws[f'E{row}'].value,
                    'monto': ws[f'F{row}'].value,
                    'notas': ws[f'G{row}'].value,
                })
        return contratos
    
    def obtener_categorias_ingresos(self):
        """Obtiene categorías únicas de ingresos"""
        ingresos = self.obtener_ingresos()
        categorias = set()
        for ingreso in ingresos:
            if ingreso['categoria']:
                categorias.add(ingreso['categoria'])
        return sorted(list(categorias))
    
    def obtener_categorias_gastos(self):
        """Obtiene categorías únicas de gastos"""
        gastos = self.obtener_gastos()
        categorias = set()
        for gasto in gastos:
            if gasto['categoria']:
                categorias.add(gasto['categoria'])
        return sorted(list(categorias))
    
    def calcular_totales_mes(self, año, mes):
        """Calcula totales de ingresos y gastos para un mes específico"""
        ingresos = self.obtener_ingresos()
        gastos = self.obtener_gastos()
        contratos = self.obtener_contratos_labor()
        
        total_ingresos = 0
        total_gastos = 0
        total_contratos = 0
        
        for ingreso in ingresos:
            if ingreso['monto'] and ingreso['fecha']:
                fecha = self._parse_fecha(ingreso['fecha'])
                if fecha and fecha.year == año and fecha.month == mes:
                    total_ingresos += float(ingreso['monto'])
        
        for gasto in gastos:
            if gasto['monto'] and gasto['fecha']:
                fecha = self._parse_fecha(gasto['fecha'])
                if fecha and fecha.year == año and fecha.month == mes:
                    total_gastos += float(gasto['monto'])
        
        for contrato in contratos:
            if contrato['monto'] and contrato['fecha']:
                fecha = self._parse_fecha(contrato['fecha'])
                if fecha and fecha.year == año and fecha.month == mes:
                    total_contratos += float(contrato['monto'])
        
        return {
            'ingresos': total_ingresos,
            'gastos': total_gastos + total_contratos,
            'contratos': total_contratos,
            'balance': total_ingresos - (total_gastos + total_contratos)
        }
    
    def _parse_fecha(self, fecha):
        """Convierte diferentes formatos de fecha a datetime"""
        if isinstance(fecha, datetime):
            return fecha
        if isinstance(fecha, str):
            for fmt in ['%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y']:
                try:
                    return datetime.strptime(fecha, fmt)
                except:
                    pass
        return None
